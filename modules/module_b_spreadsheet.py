"""
MODULE B: Human-in-the-Loop (HITL) Spreadsheet Generator
==========================================================
Writes filtered leads + AI-drafted emails to Google Sheets (primary)
and a local Excel file (backup). Approval column uses data validation
dropdown: Pending / Approved / Rejected.

Dependencies:
    pip install gspread google-auth openpyxl
"""

import os
import logging
from datetime import datetime
from dataclasses import asdict

import gspread
from google.oauth2.service_account import Credentials
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

log = logging.getLogger(__name__)

# ──────────────────────────────────────────────
# COLUMN SCHEMA  (order matters for Excel/Sheets)
# ──────────────────────────────────────────────
COLUMNS = [
    "Date Sourced",
    "Target Name",
    "Company",
    "Position",
    "Target Email",
    "Email Verified",    # TRUE = came from a lookup API; FALSE = pattern-guessed, dispatch blocked
    "Email Source",      # e.g. "hunter (confidence=85)" | "pattern-inferred (UNVERIFIED)"
    "Company Type",
    "LinkedIn URL",
    "Location",
    "Reason for Outreach",
    "Drafted Email Body",
    "Approval Status",
]

APPROVAL_OPTIONS = ["Pending", "Approved", "Rejected", "Manual-Verify"]
# "Manual-Verify" = human has confirmed the pattern-inferred email is real
# before it goes to dispatch. Module D gates on Approved AND (Email Verified=TRUE OR status=Manual-Verify).


# ──────────────────────────────────────────────
# GOOGLE SHEETS WRITER
# ──────────────────────────────────────────────
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

def get_gspread_client() -> gspread.Client:
    """Authenticates using a service-account JSON file."""
    creds_path = os.environ.get("GOOGLE_CREDENTIALS_JSON", "credentials.json")
    creds = Credentials.from_service_account_file(creds_path, scopes=SCOPES)
    return gspread.authorize(creds)


def write_to_google_sheets(
    leads_with_drafts: list[dict],
    spreadsheet_id: str,
    worksheet_name: str = "Outbound Pipeline",
) -> str:
    """
    Upserts rows to Google Sheets. Adds header if sheet is empty.
    Returns the spreadsheet URL.

    leads_with_drafts: list of dicts, each must have all COLUMNS keys.
    """
    gc = get_gspread_client()
    try:
        sh = gc.open_by_key(spreadsheet_id)
    except gspread.exceptions.SpreadsheetNotFound:
        log.error(f"Spreadsheet {spreadsheet_id} not found. Check ID and sharing permissions.")
        raise

    # Get or create worksheet
    try:
        ws = sh.worksheet(worksheet_name)
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title=worksheet_name, rows=1000, cols=len(COLUMNS))

    existing = ws.get_all_values()

    # Write header if blank
    if not existing or existing[0] != COLUMNS:
        ws.insert_row(COLUMNS, 1)
        ws.format("1:1", {
            "textFormat": {"bold": True},
            "backgroundColor": {"red": 0.12, "green": 0.24, "blue": 0.42},
        })
        # Freeze header row
        ws.freeze(rows=1)
        existing_emails: set[str] = set()
    else:
        # Build set of already-logged emails to avoid duplicates
        rows = ws.get_all_records()
        existing_emails = {r.get("Target Email", "") for r in rows}

    new_rows = []
    for lead in leads_with_drafts:
        if lead.get("Target Email", "") in existing_emails:
            log.info(f"Skipping duplicate: {lead.get('Target Email')}")
            continue
        row = [lead.get(col, "") for col in COLUMNS]
        # Default approval status
        if not row[COLUMNS.index("Approval Status")]:
            # Unverified emails get "Manual-Verify" instead of "Pending" so they
            # stand out in the sheet and can't be accidentally dispatched.
            email_verified = str(lead.get("Email Verified", "")).upper()
            row[COLUMNS.index("Approval Status")] = (
                "Manual-Verify" if email_verified in ("FALSE", "NO", "") else "Pending"
            )
        new_rows.append(row)

    if new_rows:
        ws.append_rows(new_rows, value_input_option="USER_ENTERED")
        log.info(f"Appended {len(new_rows)} rows to Google Sheets.")

        # Approval Status is now column M (index 12, 0-based = col 12)
        approval_col_idx = COLUMNS.index("Approval Status")
        last_data_row = len(existing) + len(new_rows)

        requests = [
            # Dropdown on Approval Status column
            {
                "setDataValidation": {
                    "range": {
                        "sheetId": ws.id,
                        "startRowIndex": 1,
                        "endRowIndex": last_data_row + 10,
                        "startColumnIndex": approval_col_idx,
                        "endColumnIndex": approval_col_idx + 1,
                    },
                    "rule": {
                        "condition": {
                            "type": "ONE_OF_LIST",
                            "values": [{"userEnteredValue": v} for v in APPROVAL_OPTIONS],
                        },
                        "showCustomUi": True,
                        "strict": True,
                    },
                }
            },
            # Conditional formatting — amber background for Manual-Verify rows
            # so unverified emails are visually obvious without opening each row.
            {
                "addConditionalFormatRule": {
                    "rule": {
                        "ranges": [{
                            "sheetId": ws.id,
                            "startRowIndex": 1,
                            "endRowIndex": last_data_row + 10,
                        }],
                        "booleanRule": {
                            "condition": {
                                "type": "TEXT_EQ",
                                "values": [{"userEnteredValue": "Manual-Verify"}],
                            },
                            "format": {
                                "backgroundColor": {"red": 1.0, "green": 0.85, "blue": 0.4},
                            },
                        },
                    },
                    "index": 0,
                }
            },
        ]
        sh.batch_update({"requests": requests})
    else:
        log.info("No new rows to add (all duplicates).")

    return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"


# ──────────────────────────────────────────────
# EXCEL BACKUP WRITER
# ──────────────────────────────────────────────
HEADER_FILL        = PatternFill(start_color="1E3D6A", end_color="1E3D6A", fill_type="solid")
HEADER_FONT        = Font(color="FFFFFF", bold=True, size=10)
ALT_ROW_FILL       = PatternFill(start_color="EEF2F7", end_color="EEF2F7", fill_type="solid")
UNVERIFIED_ROW_FILL = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
# Amber = pattern-inferred email, needs manual check before you can approve it

COLUMN_WIDTHS = {
    "Date Sourced":        14,
    "Target Name":         22,
    "Company":             26,
    "Position":            30,
    "Target Email":        32,
    "Email Verified":      14,
    "Email Source":        38,
    "Company Type":        14,
    "LinkedIn URL":        40,
    "Location":            22,
    "Reason for Outreach": 45,
    "Drafted Email Body":  80,
    "Approval Status":     16,
}


def write_to_excel(
    leads_with_drafts: list[dict],
    output_path: str = "output/outbound_pipeline.xlsx",
) -> str:
    """Creates or updates an Excel workbook with all leads."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Load existing or create new
    try:
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        existing_emails = {ws.cell(row=r, column=5).value for r in range(2, ws.max_row + 1)}
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Outbound Pipeline"
        existing_emails: set = set()

        # Write header
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws.column_dimensions[cell.column_letter].width = COLUMN_WIDTHS.get(col_name, 20)

        ws.row_dimensions[1].height = 25
        ws.freeze_panes = "A2"

    # Dropdown validation for Approval Status — now column M (index 13, letter M)
    approval_col_letter = chr(ord("A") + COLUMNS.index("Approval Status"))
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(APPROVAL_OPTIONS)}"',
        showDropDown=False,
        sqref=f"{approval_col_letter}2:{approval_col_letter}5000",
    )
    ws.add_data_validation(dv)

    start_row = ws.max_row + 1

    for i, lead in enumerate(leads_with_drafts):
        email = lead.get("Target Email", "")
        if email in existing_emails:
            continue

        row_num = start_row + i
        email_verified = str(lead.get("Email Verified", "")).upper() in ("TRUE", "YES", "1")

        # Unverified rows get amber fill so they stand out immediately.
        # Verified rows get standard alt-row banding.
        if not email_verified:
            fill = UNVERIFIED_ROW_FILL
        else:
            fill = ALT_ROW_FILL if row_num % 2 == 0 else None

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value = lead.get(col_name, "")
            if col_name == "Approval Status" and not value:
                value = "Manual-Verify" if not email_verified else "Pending"
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

        ws.row_dimensions[row_num].height = 60
        existing_emails.add(email)

    wb.save(output_path)
    log.info(f"Excel written to {output_path}")
    return output_path


# ──────────────────────────────────────────────
# COMBINED WRITER — calls both
# ──────────────────────────────────────────────
def write_leads(
    leads_with_drafts: list[dict],
    spreadsheet_id: str | None = None,
    excel_path: str = "output/outbound_pipeline.xlsx",
):
    """
    Writes to both Google Sheets and local Excel.
    Pass spreadsheet_id=None to skip Google Sheets.
    """
    if spreadsheet_id:
        try:
            url = write_to_google_sheets(leads_with_drafts, spreadsheet_id)
            log.info(f"Google Sheets updated: {url}")
        except Exception as e:
            log.error(f"Google Sheets write failed: {e}. Falling back to Excel only.")

    xlsx_path = write_to_excel(leads_with_drafts, excel_path)
    return xlsx_path

# Orchestrator-compatible aliases
def append_leads(leads, spreadsheet_id=None, excel_path="output/outbound_pipeline.xlsx"):
    return write_leads(leads, spreadsheet_id=spreadsheet_id, excel_path=excel_path)

def export_excel(leads, output_path="output/outbound_pipeline.xlsx"):
    return write_to_excel(leads, output_path=output_path)

def get_stats(spreadsheet_id, worksheet_name="Outbound Pipeline"):
    gc = get_gspread_client()
    ws = gc.open_by_key(spreadsheet_id).worksheet(worksheet_name)
    rows = ws.get_all_records()
    total = len(rows)
    approved = sum(1 for r in rows if r.get("Approval Status","").lower() == "approved")
    pending  = sum(1 for r in rows if r.get("Approval Status","").lower() == "pending")
    sent     = sum(1 for r in rows if "sent" in r.get("Approval Status","").lower())
    unverified = sum(1 for r in rows if str(r.get("Email Verified","")).upper() not in ("TRUE","YES","1"))
    return {"total": total, "approved": approved, "pending": pending,
            "sent": sent, "unverified_emails": unverified}
